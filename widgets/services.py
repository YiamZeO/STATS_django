import datetime
import logging
from copy import copy

from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook

from djangoProject.settings import clickhouse_default_client
from utils.utils import Utils, ResponseObject
from widgets.models import DegWidget, DegTypes, DegField


class DegDataService:
    """
    Сервис для раздела ДЭГ
    """

    def __init__(self, client=clickhouse_default_client):
        self.__client = client

    def update_deg_widgets_collection(self, schema):
        """
        Обновление коллекции DegWidget данными из ClickHouse
        :param schema: схема в ClickHouse
        :return: список добавленных DegWidget
        """

        deg_widgets = DegWidget.objects()
        deg_widget_max_order = max(deg_widgets, key=lambda deg_widget: deg_widget.order,
                                   default=DegWidget(order=1)).order
        deg_widgets = [(deg_widget.table_name, deg_widget.type) for deg_widget in deg_widgets]
        current_tables = Utils.get_tables_comments(schema, self.__client)
        new_deg_widgets = []
        for table in current_tables:
            try:
                if (table, DegTypes.EXPORT) not in deg_widgets or (table, DegTypes.BOARD) not in deg_widgets:
                    deg_widget_fields = list()
                    current_columns = Utils.get_columns_comments(f'{schema}.{table}', self.__client)
                    for order, column in enumerate(current_columns, 1):
                        deg_field = DegField(name=column,
                                             russian_name=current_columns[column],
                                             show=True,
                                             order=order
                                             )
                        deg_widget_fields.append(deg_field)
                    deg_widget = DegWidget(
                        alias=table.replace('_', ''),
                        schema=schema,
                        order=deg_widget_max_order,
                        russian_name=current_tables[table],
                        show=True,
                        table_name=table,
                        fields_list=deg_widget_fields
                    )
                    if (table, DegTypes.EXPORT) not in deg_widgets:
                        deg_widget_ = copy(deg_widget)
                        deg_widget_.type = DegTypes.EXPORT
                        new_deg_widgets.append(deg_widget_)
                    if (table, DegTypes.BOARD) not in deg_widgets:
                        deg_widget_ = copy(deg_widget)
                        deg_widget_.type = DegTypes.BOARD
                        new_deg_widgets.append(deg_widget_)
                    deg_widget_max_order += 1
            except Exception as e:
                logging.error(f'Table {table} not proceed. Error: {e}')
        if len(new_deg_widgets) > 0:
            return DegWidget.objects().insert(new_deg_widgets)
        else:
            return list()

    @staticmethod
    def __get_meta(schema, alias):
        """
        Формирование мета информации о DegWidget
        :param schema: схема в ClickHouse
        :param alias: алиас DegWidget
        :return: мета информация о DegWidget
        """

        deg_widget = DegWidget.objects(alias=alias, schema=schema).first()
        meta_info = {field.name: field.russian_name for field in deg_widget.fields_list}
        meta_info['table_alias'] = deg_widget.alias
        if deg_widget.russian_name:
            meta_info['table_russian_name'] = deg_widget.russian_name
        return meta_info

    def get_data(self, schema, alias, deg_type, date_from, date_to, extractor_code=None):
        """
        Получение данных DegWidget из ClickHouse
        :param schema: схема в ClickHouse
        :param alias: алиас DegWidget
        :param deg_type: тип DegWidget
        :param date_from: дата начала промежутка
        :param date_to: дата конца промежутка
        :param extractor_code: способ обработки и выгрузки данных
        :return: данные DegWidget из ClickHouse
        """

        deg_widget = DegWidget.objects(alias=alias, schema=schema, type=deg_type).first()
        if not deg_widget:
            raise Exception(f'DegWidget: schema = {schema}, alias = {alias} -- not exists')
        table_name = f'{schema}.{deg_widget.table_name}'
        sql = (f'{self.__select_with_fields(deg_widget.fields_list)} from {table_name}'
               f'{' where date between {date_from: Date} and {date_to: Date}' if date_from and date_to else ''} order by date')
        return ResponseObject(
            self.__deg_data_extraction(sql, date_from, date_to, extractor_code),
            self.__get_meta(schema, alias)
        )

    def __deg_data_extraction(self, sql, date_from, date_to, extractor_code):
        """
        Получение данных DegWidget из ClickHouse способом, зависящим от extractor_code
        :param sql: зарос для получения данных
        :param date_from: дата начала промежутка
        :param date_to: дата конца промежутка
        :param extractor_code: способ обработки и выгрузки данных
        :return: данные DegWidget из ClickHouse способом, зависящим от extractor_code
        """

        if extractor_code == 'for_details':
            with self.__client.query_rows_stream(sql, parameters={
                'date_from': date_from,
                'date_to': date_to
            }) as stream:
                data = {column_name: list() for column_name in stream.source.column_names if column_name != 'date'}
                for row in stream:
                    for i, field_value in enumerate(row):
                        if stream.source.column_names[i] != 'date':
                            data[stream.source.column_names[i]].append({
                                'date': row[stream.source.column_names.index('date')],
                                'value': field_value
                            })
                return data
        else:
            with self.__client.query_rows_stream(sql, parameters={
                'date_from': date_from,
                'date_to': date_to
            }) as stream:
                data = [{stream.source.column_names[i]: field_value for i, field_value in enumerate(row)} for row in stream]
                return data

    @staticmethod
    def __select_with_fields(fields):
        """
        Формирование sql запроса, в зависимости от fields
        :param fields: список полей для sql запроса
        :return: sql запрос, в зависимости от fields
        """

        sql = 'select '
        if len(fields) == 0:
            sql += '*'
        else:
            sql += fields[0].name
            if len(fields) > 1:
                str_fields = [f', {field.name}' for field in fields[1:]]
                sql += ''.join(str_fields)
        return sql

    def get_deg_report(self, date_from, date_to, schema):
        """
        Получение xlsx отчета из ClickHouse данных DegWidget
        :param date_from: дата начала промежутка
        :param date_to: дата конца промежутка
        :param schema: схема в ClickHouse
        :return: xlsx отчет из ClickHouse данных DegWidget
        """

        wb = Workbook()
        ws = wb.active
        ws.title = 'ДЭГ таблицы'
        table_name_font = Font(size=16, bold=True)
        header_font = Font(size=12, bold=True)
        value_font = Font(size=12, bold=False)
        wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        deg_widgets = DegWidget.objects(type=DegTypes.EXPORT, schema=schema) if schema \
            else DegWidget.objects(type=DegTypes.EXPORT)
        row_num = 1
        for deg_widget in deg_widgets:
            cell = ws.cell(row=row_num, column=2)
            cell.font = table_name_font
            cell.value = deg_widget.russian_name
            # cell.alignment = wrap_alignment
            row_num += 1
            for column_num, field in enumerate(deg_widget.fields_list, 2):
                cell = ws.cell(row=row_num, column=column_num)
                cell.font = header_font
                cell.value = field.russian_name
                cell.alignment = wrap_alignment
                ws.column_dimensions[cell.column_letter].width = 20
            row_num += 1
            deg_widget_data = self.get_data(deg_widget.schema, deg_widget.alias, DegTypes.EXPORT, date_from, date_to)
            data_start_row = row_num
            for row_index, deg_row_data in enumerate(deg_widget_data.data, 1):
                index_cell = ws.cell(row=row_num, column=1)
                index_cell.font = header_font
                index_cell.alignment = wrap_alignment
                index_cell.value = row_index
                for column_num, field in enumerate(deg_widget.fields_list, 2):
                    cell = ws.cell(row=row_num, column=column_num)
                    cell.font = value_font
                    deg_cell_value = deg_row_data.get(field.name)
                    if isinstance(deg_cell_value, float):
                        deg_cell_value = round(deg_cell_value, 2)
                    if isinstance(deg_cell_value, datetime.date):
                        deg_cell_value = deg_cell_value.strftime('%d.%m.%Y')
                    cell.value = deg_cell_value
                    cell.alignment = wrap_alignment
                row_num += 1
            ws.row_dimensions.group(data_start_row, row_num - 1, hidden=False)
            row_num += 2
        return wb
